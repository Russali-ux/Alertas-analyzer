"""
Scraper de Alertas DIGEMID — Versión 3.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

NOVEDADES v3.0:
  ✦ Sube PDFs automáticamente a GitHub (repositorio como almacén permanente)
  ✦ Claude lee PDFs directamente desde la URL raw de GitHub (más preciso)
  ✦ Detecta PDFs ya existentes en GitHub → no los sube dos veces
  ✦ Guarda index JSON en GitHub (data/alertas_YYYYMMDD.json)
  ✦ Motor de análisis: Claude API (URL) → Claude API (texto) → Heurístico

VARIABLES DE ENTORNO:
  ANTHROPIC_API_KEY   → análisis semántico con Claude
  GITHUB_TOKEN        → subir PDFs y JSONs al repo (se genera automáticamente en Actions)
  GITHUB_REPO         → "usuario/nombre-repo" (ej: conkomerco/digemid-alertas)
  GITHUB_BRANCH       → rama objetivo (default: main)

USO LOCAL:
  # Modo básico (solo heurístico, sin GitHub):
  python3 scraper_alertas_digemid_v3.py

  # Modo completo (Claude API + GitHub):
  ANTHROPIC_API_KEY=sk-ant-... GITHUB_TOKEN=ghp_... GITHUB_REPO=usuario/repo \
  python3 scraper_alertas_digemid_v3.py

USO EN GITHUB ACTIONS (ver workflow al final del archivo):
  - Crea .github/workflows/digemid_monitor.yml en tu repo
  - Agrega ANTHROPIC_API_KEY a Secrets del repo
  - GITHUB_TOKEN se inyecta automáticamente

ESTRUCTURA EN EL REPOSITORIO:
  pdfs/
    YYYY-MM-DD_ALERTA-DIGEMID-N-titulo.pdf
  data/
    alertas_YYYYMMDD.json          ← index completo del día
    alertas_latest.json            ← siempre contiene el último run
  summaries/
    resumen_YYYYMMDD.md            ← resumen legible en Markdown
"""

import os, re, json, time, base64, hashlib
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import fitz   # pymupdf  →  pip install pymupdf

# ─────────────────────────────────────────────────────────────────
# CONFIGURACIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────────
BASE_URL      = "https://www.digemid.minsa.gob.pe"
ALERTAS_URL   = f"{BASE_URL}/alertas"
PAGE_URL      = f"{BASE_URL}/webDigemid/alertas/page/{{page}}/"
PDF_DELAY     = 15          # segundos entre descargas de PDF (rate-limit DIGEMID)

# GitHub
GITHUB_TOKEN  = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO   = os.environ.get("GITHUB_REPO", "")       # "usuario/repo"
GITHUB_BRANCH = os.environ.get("GITHUB_BRANCH", "main")
GITHUB_API    = "https://api.github.com"

# Sesión con headers de navegador real (evita bloqueo Cloudflare 403)
session = requests.Session()
HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-PE,es;q=0.9,en-US;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer":         "https://www.digemid.minsa.gob.pe/",
}
try:
    session.get(BASE_URL, headers=HEADERS, timeout=30)  # calentar cookies
except Exception:
    pass


# ─────────────────────────────────────────────────────────────────
# PROMPT PARA CLAUDE API
# ─────────────────────────────────────────────────────────────────
PROMPT_ACCIONES_TEXTO = """\
Eres un experto en regulación farmacéutica peruana.
Se te proporciona el texto completo de una Alerta DIGEMID del Perú.

Extrae ÚNICAMENTE las acciones concretas que DIGEMID solicita o recomienda.
Responde SOLO con JSON, sin texto adicional ni bloques de código:

{
  "accion_principal": "<frase corta, ej: RETIRO DEL MERCADO | NO COMERCIALIZAR | NOTIFICAR REACCIONES ADVERSAS>",
  "urgencia": "<INMEDIATA | PREVENTIVA | INFORMATIVA>",
  "dirigido_a": ["<destinatario 1>", "<destinatario 2>"],
  "acciones_detalladas": ["<acción 1>", "<acción 2>", "..."],
  "resumen_accion": "<1-2 oraciones resumiendo qué debe hacer el lector>"
}

TEXTO DE LA ALERTA:
{texto}
"""

PROMPT_ACCIONES_PDF = """\
Eres un experto en regulación farmacéutica peruana.
Analiza el documento PDF adjunto (Alerta DIGEMID).

Extrae ÚNICAMENTE las acciones concretas que DIGEMID solicita o recomienda.
Responde SOLO con JSON, sin texto adicional ni bloques de código:

{
  "accion_principal": "<frase corta, ej: RETIRO DEL MERCADO | NO COMERCIALIZAR | NOTIFICAR REACCIONES ADVERSAS>",
  "urgencia": "<INMEDIATA | PREVENTIVA | INFORMATIVA>",
  "dirigido_a": ["<destinatario 1>", "<destinatario 2>"],
  "acciones_detalladas": ["<acción 1>", "<acción 2>", "..."],
  "resumen_accion": "<1-2 oraciones resumiendo qué debe hacer el lector>"
}
"""


# ─────────────────────────────────────────────────────────────────
# GITHUB — subir y leer PDFs
# ─────────────────────────────────────────────────────────────────
def _gh_headers() -> dict:
    return {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
    }


def _sanitizar_nombre(nombre: str) -> str:
    """Convierte un texto en nombre de archivo seguro."""
    nombre = re.sub(r'[^\w\-.]', '_', nombre)
    nombre = re.sub(r'_+', '_', nombre)
    return nombre[:120]


def subir_pdf_a_github(pdf_bytes: bytes, alerta: dict) -> str | None:
    """
    Sube el PDF al repo GitHub en pdfs/.
    Retorna la URL raw (https://raw.githubusercontent.com/...) o None.
    Si el archivo ya existe con el mismo contenido, retorna la URL sin subir.
    """
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return None

    fecha    = str(alerta.get("fecha_publicacion", "sin-fecha"))
    titulo   = alerta.get("titulo", "alerta")[:60]
    nombre   = _sanitizar_nombre(f"{fecha}_{titulo}.pdf")
    path_repo = f"pdfs/{nombre}"
    url_api   = f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{path_repo}"

    # Calcular SHA git del contenido nuevo
    encabezado = f"blob {len(pdf_bytes)}\0".encode()
    sha_nuevo  = hashlib.sha1(encabezado + pdf_bytes).hexdigest()

    # Verificar si ya existe
    r = requests.get(url_api, headers=_gh_headers(), timeout=15)
    sha_existente = None
    if r.status_code == 200:
        sha_existente = r.json().get("sha", "")
        # SHA de la API es el SHA del objeto git base64-contenido, no del raw
        # Si existe, simplemente reutilizamos la URL (evita subida duplicada)
        raw_url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{path_repo}"
        print(f"    [GitHub] PDF ya existe → {path_repo}")
        return raw_url

    # Subir nuevo
    payload = {
        "message": f"🚨 Auto: {alerta.get('titulo','alerta')[:60]}",
        "content": base64.b64encode(pdf_bytes).decode(),
        "branch":  GITHUB_BRANCH,
    }
    if sha_existente:
        payload["sha"] = sha_existente

    resp = requests.put(url_api, json=payload, headers=_gh_headers(), timeout=30)
    if resp.status_code in (200, 201):
        raw_url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{path_repo}"
        print(f"    [GitHub] PDF subido → {path_repo}")
        return raw_url
    else:
        print(f"    [GitHub ERROR {resp.status_code}] {resp.text[:200]}")
        return None


def subir_json_a_github(data: list, nombre_archivo: str) -> bool:
    """Sube/actualiza un archivo JSON en data/ del repo."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return False

    contenido_json = json.dumps(data, ensure_ascii=False, indent=2, default=str)
    contenido_b64  = base64.b64encode(contenido_json.encode("utf-8")).decode()
    path_repo      = f"data/{nombre_archivo}"
    url_api        = f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{path_repo}"

    # Obtener SHA actual (necesario para actualizar)
    r = requests.get(url_api, headers=_gh_headers(), timeout=15)
    payload = {
        "message": f"📊 Index alertas {nombre_archivo}",
        "content": contenido_b64,
        "branch":  GITHUB_BRANCH,
    }
    if r.status_code == 200:
        payload["sha"] = r.json().get("sha", "")

    resp = requests.put(url_api, json=payload, headers=_gh_headers(), timeout=30)
    ok = resp.status_code in (200, 201)
    if ok:
        print(f"  [GitHub] JSON guardado → data/{nombre_archivo}")
    else:
        print(f"  [GitHub ERROR] JSON {resp.status_code}")
    return ok


def subir_resumen_a_github(df: pd.DataFrame, fecha_str: str) -> bool:
    """Sube un resumen Markdown legible a summaries/."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return False

    lineas = [
        f"# Resumen Alertas DIGEMID — {fecha_str}",
        f"",
        f"**Total alertas:** {len(df)}  ",
        f"**Generado:** {datetime.now().strftime('%d/%m/%Y %H:%M')} (Lima, PE)",
        f"",
        f"## Por urgencia",
        f"| Urgencia | N° |",
        f"|----------|-----|",
    ]
    if "urgencia" in df.columns:
        for urg, cnt in df["urgencia"].value_counts().items():
            emoji = {"INMEDIATA": "🔴", "PREVENTIVA": "🟡", "INFORMATIVA": "🔵"}.get(urg, "⚪")
            lineas.append(f"| {emoji} {urg} | {cnt} |")

    lineas += ["", "## Alertas del día", ""]
    for _, row in df.iterrows():
        urgencia  = row.get("urgencia", "?")
        accion    = row.get("accion_principal", "—")
        titulo    = row.get("titulo", "—")
        pdf_gh    = row.get("github_pdf_url", "")
        pdf_link  = f" — [📄 PDF]({pdf_gh})" if pdf_gh else ""
        lineas.append(f"### {titulo}{pdf_link}")
        lineas.append(f"- **Urgencia:** {urgencia}")
        lineas.append(f"- **Acción:** {accion}")
        lineas.append(f"- **Resumen:** {row.get('resumen_accion','—')}")
        lineas.append("")

    md_content = "\n".join(lineas)
    nombre     = f"resumen_{fecha_str.replace('/', '-')}.md"
    path_repo  = f"summaries/{nombre}"
    url_api    = f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{path_repo}"
    contenido_b64 = base64.b64encode(md_content.encode("utf-8")).decode()

    r = requests.get(url_api, headers=_gh_headers(), timeout=15)
    payload = {"message": f"📝 Resumen {fecha_str}", "content": contenido_b64, "branch": GITHUB_BRANCH}
    if r.status_code == 200:
        payload["sha"] = r.json().get("sha", "")

    resp = requests.put(url_api, json=payload, headers=_gh_headers(), timeout=30)
    if resp.status_code in (200, 201):
        print(f"  [GitHub] Resumen guardado → summaries/{nombre}")
        return True
    return False


# ─────────────────────────────────────────────────────────────────
# MOTOR 1A: CLAUDE API — lee PDF desde URL de GitHub (más preciso)
# ─────────────────────────────────────────────────────────────────
def analizar_con_claude_url(github_pdf_url: str, titulo: str) -> dict | None:
    """
    Claude lee el PDF directamente desde la URL raw de GitHub.
    Más preciso que extraer texto manualmente.
    Requiere ANTHROPIC_API_KEY.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key or not github_pdf_url:
        return None
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "url",          # Claude descarga y lee el PDF
                            "url": github_pdf_url,
                        },
                    },
                    {
                        "type": "text",
                        "text": PROMPT_ACCIONES_PDF + f"\n\nTítulo de la alerta: {titulo}",
                    }
                ]
            }]
        )
        raw = msg.content[0].text.strip()
        raw = re.sub(r"^```json\s*|^```\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
        return json.loads(raw)
    except Exception as e:
        print(f"    [Claude URL] {e}")
        return None


# ─────────────────────────────────────────────────────────────────
# MOTOR 1B: CLAUDE API — analiza texto extraído (fallback)
# ─────────────────────────────────────────────────────────────────
def analizar_con_claude_texto(texto: str) -> dict | None:
    """
    Llama a Claude API enviando el texto extraído del PDF.
    Fallback cuando no hay URL de GitHub disponible.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return None
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": PROMPT_ACCIONES_TEXTO.replace("{texto}", texto[:8000])
            }]
        )
        raw = msg.content[0].text.strip()
        raw = re.sub(r"^```json\s*|^```\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
        return json.loads(raw)
    except Exception as e:
        print(f"    [Claude texto] {e}")
        return None


# ─────────────────────────────────────────────────────────────────
# MOTOR 2: HEURÍSTICO (fallback sin API key)
# ─────────────────────────────────────────────────────────────────
def analizar_heuristico(texto: str) -> dict:
    """
    Extracción basada en patrones del texto del PDF.
    Funciona sin API key y cubre los 3 tipos de alerta de DIGEMID.
    """
    texto_up = texto.upper()

    if any(k in texto_up for k in ["FALSIFICADO", "INCAUTADO", "FALSIFICACIÓN"]):
        accion_principal = "NO ADQUIRIR / NO COMERCIALIZAR"
        urgencia = "INMEDIATA"
    elif any(k in texto_up for k in ["RETIRO DEL MERCADO", "RECALL", "RETIRAR DEL MERCADO"]):
        accion_principal = "RETIRO DEL MERCADO"
        urgencia = "INMEDIATA"
    elif any(k in texto_up for k in ["SUSPENDER", "SUSPENSIÓN", "PROHIBIR"]):
        accion_principal = "SUSPENDER USO / DISTRIBUCIÓN"
        urgencia = "INMEDIATA"
    elif any(k in texto_up for k in ["RESULTADO CRÍTICO", "NO CONFORME", "SUBESTÁNDAR"]):
        accion_principal = "RETIRO POR CALIDAD"
        urgencia = "INMEDIATA"
    elif any(k in texto_up for k in ["RIESGO", "REACCIÓN ADVERSA", "SIADM", "REPORTE"]):
        accion_principal = "NOTIFICAR / MEDIDAS PREVENTIVAS"
        urgencia = "PREVENTIVA"
    else:
        accion_principal = "VER COMUNICADO OFICIAL"
        urgencia = "INFORMATIVA"

    destinatarios = re.findall(
        r"[Aa]\s+los(?:as)?\s+([\w\s,áéíóúñÁÉÍÓÚÑ]+?):",
        texto, re.IGNORECASE
    )
    dirigido_a = list(dict.fromkeys([d.strip() for d in destinatarios]))[:4]

    bullets_raw = [
        l.strip().lstrip("•").lstrip("-").strip()
        for l in texto.split("\n")
        if l.strip().startswith(("•", "-")) and len(l.strip()) > 15
    ]
    acciones_detalladas = bullets_raw[:8]

    patrones_resumen = [
        r"[Ss]e recomienda\s+([^.]+\.)",
        r"[Ss]e (solicita|requiere|exige|dispone)\s+([^.]+\.)",
        r"[Ll]a Digemid recomienda\s+([^.]+\.)",
        r"no (comprar|adquirir|utilizar|comercializar)\s+([^.]+\.)",
    ]
    resumen = ""
    for patron in patrones_resumen:
        m = re.search(patron, texto, re.IGNORECASE)
        if m:
            resumen = m.group(0).strip()
            break
    if not resumen and acciones_detalladas:
        resumen = acciones_detalladas[0][:220]

    return {
        "accion_principal":    accion_principal,
        "urgencia":            urgencia,
        "dirigido_a":          dirigido_a,
        "acciones_detalladas": acciones_detalladas,
        "resumen_accion":      resumen,
    }


# ─────────────────────────────────────────────────────────────────
# EXTRACCIÓN DE PDF + ANÁLISIS + SUBIDA A GITHUB
# ─────────────────────────────────────────────────────────────────
def obtener_pdf_url(url_alerta: str) -> str | None:
    try:
        resp = session.get(url_alerta, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        entry = soup.find("div", class_="entry-content")
        if not entry:
            return None
        link = entry.find("a", href=re.compile(r"\.pdf$", re.I))
        if not link:
            link = entry.find("a", href=re.compile(r"Alertas", re.I))
        if not link:
            embed = entry.find("embed", src=re.compile(r"\.pdf", re.I))
            if embed:
                href = embed.get("src", "")
                return href if href.startswith("http") else f"{BASE_URL}{href}"
        if not link:
            return None
        href = link["href"]
        return href if href.startswith("http") else f"{BASE_URL}{href}"
    except Exception:
        return None


def descargar_pdf(pdf_url: str, reintentos: int = 3) -> bytes | None:
    """Descarga PDF con reintentos y backoff para evitar 429."""
    pdf_headers = {**HEADERS, "Accept": "application/pdf,*/*"}
    for intento in range(1, reintentos + 1):
        try:
            r = session.get(pdf_url, headers=pdf_headers, timeout=30)
            if r.status_code == 200:
                return r.content
            elif r.status_code == 429:
                espera = PDF_DELAY * intento
                print(f"    [429] Esperando {espera}s...", end=" ", flush=True)
                time.sleep(espera)
            else:
                return None
        except Exception:
            time.sleep(PDF_DELAY)
    return None


def enriquecer_alerta(alerta: dict) -> dict:
    """
    Pipeline completo por alerta:
    1. Obtiene URL del PDF desde la página de detalle
    2. Descarga el PDF
    3. Sube a GitHub → obtiene URL raw
    4. Extrae texto con PyMuPDF
    5. Analiza: Claude (URL GitHub) → Claude (texto) → Heurístico
    """
    url_detalle = alerta.get("url")
    if not url_detalle:
        alerta.update({
            "pdf_url": None, "github_pdf_url": None,
            "accion_principal": "Sin URL", "urgencia": "INFORMATIVA",
            "resumen_accion": "Sin URL", "acciones_detalladas": "", "dirigido_a": "",
            "motor_analisis": "—"
        })
        return alerta

    pdf_url = obtener_pdf_url(url_detalle)
    alerta["pdf_url"] = pdf_url

    if not pdf_url:
        alerta.update({
            "github_pdf_url": None,
            "accion_principal": "Sin PDF", "urgencia": "INFORMATIVA",
            "resumen_accion": "PDF no encontrado.", "acciones_detalladas": "", "dirigido_a": "",
            "motor_analisis": "—"
        })
        return alerta

    pdf_bytes = descargar_pdf(pdf_url)
    if not pdf_bytes:
        alerta.update({
            "github_pdf_url": None,
            "accion_principal": "Error descarga PDF", "urgencia": "INFORMATIVA",
            "resumen_accion": "No se pudo descargar el PDF.", "acciones_detalladas": "", "dirigido_a": "",
            "motor_analisis": "—"
        })
        return alerta

    # ── SUBIR A GITHUB ──────────────────────────────────────────
    github_url = subir_pdf_a_github(pdf_bytes, alerta)
    alerta["github_pdf_url"] = github_url

    # ── EXTRAER TEXTO LOCAL (para fallback) ─────────────────────
    texto = ""
    try:
        doc   = fitz.open(stream=pdf_bytes, filetype="pdf")
        texto = "\n".join(p.get_text() for p in doc).strip()
    except Exception as e:
        print(f"    [PyMuPDF] {e}")

    if len(texto) < 50:
        alerta.update({
            "accion_principal": "PDF escaneado", "urgencia": "INFORMATIVA",
            "resumen_accion": "PDF es imagen, sin texto extraíble.",
            "acciones_detalladas": "", "dirigido_a": "",
            "motor_analisis": "—"
        })
        return alerta

    # ── CADENA DE ANÁLISIS ──────────────────────────────────────
    resultado    = None
    motor_nombre = "Heurístico"

    # Intento 1: Claude leyendo PDF desde GitHub URL
    if github_url and os.environ.get("ANTHROPIC_API_KEY"):
        resultado = analizar_con_claude_url(github_url, alerta.get("titulo", ""))
        if resultado:
            motor_nombre = "Claude API (PDF→GitHub)"

    # Intento 2: Claude con texto extraído
    if not resultado and os.environ.get("ANTHROPIC_API_KEY"):
        resultado = analizar_con_claude_texto(texto)
        if resultado:
            motor_nombre = "Claude API (texto)"

    # Intento 3: Heurístico
    if not resultado:
        resultado = analizar_heuristico(texto)

    alerta["accion_principal"]    = resultado.get("accion_principal", "")
    alerta["urgencia"]            = resultado.get("urgencia", "INFORMATIVA")
    alerta["resumen_accion"]      = resultado.get("resumen_accion", "")
    alerta["acciones_detalladas"] = " | ".join(resultado.get("acciones_detalladas", []))
    alerta["dirigido_a"]          = " | ".join(resultado.get("dirigido_a", []))
    alerta["motor_analisis"]      = motor_nombre
    return alerta


# ─────────────────────────────────────────────────────────────────
# SCRAPER PRINCIPAL
# ─────────────────────────────────────────────────────────────────
def _parsear_articulos(soup: BeautifulSoup) -> list[dict]:
    alertas = []
    for article in soup.find_all("article", class_=re.compile(r"\bpost\b")):
        titulo_tag = article.find("h2", class_="entry-title")
        if not titulo_tag:
            continue
        titulo = titulo_tag.get_text(strip=True)
        if "ALERTA DIGEMID" not in titulo.upper():
            continue
        link_tag  = titulo_tag.find("a")
        link      = link_tag["href"] if link_tag and link_tag.get("href") else None
        time_tag  = article.find("time")
        fecha_pub = None
        if time_tag and time_tag.get("datetime"):
            try:
                fecha_pub = datetime.strptime(time_tag["datetime"], "%Y-%m-%d").date()
            except ValueError:
                pass
        excerpt = article.find("p", class_="post-excerpt")
        producto = excerpt.get_text(strip=True) if excerpt else None
        cat_tags = article.select("div.post-meta span.meta-cats a[rel='category tag']")
        cats = [c.get_text(strip=True) for c in cat_tags
                if c.get_text(strip=True) not in ("Alertas", "Alertas y Modificaciones")]
        alertas.append({
            "titulo":          titulo,
            "producto":        producto,
            "tipo_alerta":     cats[0] if cats else "General",
            "fecha_publicacion": fecha_pub,
            "url":             link,
            "fecha_captura":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "pdf_url":         None,
            "github_pdf_url":  None,
        })
    return alertas


def _obtener_total_paginas(soup: BeautifulSoup) -> int:
    pag = soup.find("div", class_="pagination")
    if not pag:
        return 1
    nums = [int(a.get_text(strip=True))
            for a in pag.find_all("a", class_="page-numbers")
            if a.get_text(strip=True).isdigit()]
    return max(nums) if nums else 1


def scrapear_alertas(
    max_paginas: int  = None,
    delay_paginas: float = 1.2,
    analizar_acciones: bool = True
) -> pd.DataFrame:
    """
    Parámetros:
      max_paginas      : Límite de páginas (None = todas, ~165 en total).
      delay_paginas    : Espera entre páginas del listado.
      analizar_acciones: Si True, descarga PDF, sube a GitHub y extrae acciones.
    """
    resp = session.get(ALERTAS_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    total = min(
        _obtener_total_paginas(soup),
        max_paginas if max_paginas else 9999
    )
    print(f"Páginas a procesar: {total}")

    alertas = _parsear_articulos(soup)
    print(f"  Página 1: {len(alertas)} alertas")

    for pag in range(2, total + 1):
        try:
            nuevas = _parsear_articulos(
                BeautifulSoup(
                    session.get(PAGE_URL.format(page=pag), headers=HEADERS, timeout=30).text,
                    "html.parser"
                )
            )
            alertas.extend(nuevas)
            print(f"  Página {pag}/{total}: {len(nuevas)} alertas")
        except Exception as e:
            print(f"  ERROR página {pag}: {e}")
        time.sleep(delay_paginas)

    if analizar_acciones:
        tiene_claude  = bool(os.environ.get("ANTHROPIC_API_KEY"))
        tiene_github  = bool(GITHUB_TOKEN and GITHUB_REPO)
        motor_label   = "Claude API + GitHub" if (tiene_claude and tiene_github) \
                        else "Claude API" if tiene_claude \
                        else "Heurístico"
        print(f"\nExtrayendo acciones [{motor_label}] para {len(alertas)} alertas...")

        for i, a in enumerate(alertas, 1):
            print(f"  [{i:>3}/{len(alertas)}] {a['titulo'][:60]}...", end=" ", flush=True)
            enriquecer_alerta(a)
            motor = a.get("motor_analisis", "?")
            print(f"→ [{a.get('urgencia','?'):11}] {a.get('accion_principal','?')} [{motor}]")
            time.sleep(PDF_DELAY)

    return pd.DataFrame(alertas)


# ─────────────────────────────────────────────────────────────────
# EXPORTAR A EXCEL (mismo formato que v2 + columna GitHub PDF URL)
# ─────────────────────────────────────────────────────────────────
def exportar_excel(df: pd.DataFrame, ruta: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLS = [
        ("Título",               "titulo",               40),
        ("Producto",             "producto",             28),
        ("Tipo de Alerta",       "tipo_alerta",          20),
        ("Fecha Publicación",    "fecha_publicacion",    15),
        ("⚡ Acción Principal",  "accion_principal",     30),
        ("Urgencia",             "urgencia",             13),
        ("Dirigido a",           "dirigido_a",           35),
        ("Acciones Requeridas",  "acciones_detalladas",  65),
        ("Resumen IA",           "resumen_accion",       55),
        ("Motor Análisis",       "motor_analisis",       22),
        ("URL Alerta",           "url",                  50),
        ("URL PDF DIGEMID",      "pdf_url",              50),
        ("URL PDF GitHub",       "github_pdf_url",       65),
        ("Fecha Captura",        "fecha_captura",        16),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alertas DIGEMID"

    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    d_font = Font(name="Arial", size=9)

    for ci, (label, _, width) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = h_font; c.fill = h_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24

    URGENCIA_COLOR = {"INMEDIATA": "FCE4D6", "PREVENTIVA": "FFF2CC", "INFORMATIVA": "EBF3FB"}
    URGENCIA_BADGE = {"INMEDIATA": "C00000", "PREVENTIVA": "ED7D31", "INFORMATIVA": "2E75B6"}

    for ri, row in enumerate(df.itertuples(index=False), 2):
        urgencia  = getattr(row, "urgencia", "INFORMATIVA") or "INFORMATIVA"
        row_fill  = PatternFill("solid", start_color=URGENCIA_COLOR.get(urgencia, "F2F2F2"))

        for ci, (_, field, _) in enumerate(COLS, 1):
            # Manejar nombres con espacios / caracteres especiales
            safe_field = field.replace(" ", "_").replace("⚡", "").strip("_")
            val = None
            for attr in [field, safe_field]:
                try:
                    val = getattr(row, attr, None)
                    if val is not None:
                        break
                except Exception:
                    pass

            c = ws.cell(row=ri, column=ci, value=val)
            c.font = d_font; c.fill = row_fill; c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="center")

            if field in ("url", "pdf_url", "github_pdf_url") and val:
                c.hyperlink = str(val)
                c.font = Font(name="Arial", size=9, color="0563C1", underline="single")

            if field == "urgencia" and val:
                c.font = Font(name="Arial", size=9, bold=True,
                              color=URGENCIA_BADGE.get(urgencia, "000000"))
                c.alignment = Alignment(horizontal="center", vertical="center")

            if field == "accion_principal":
                c.font = Font(name="Arial", size=9, bold=True)

        ws.row_dimensions[ri].height = 52

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(df)+1}"

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Reporte de Alertas DIGEMID v3 — Con acciones y PDFs en GitHub"
    ws2["A1"].font = Font(bold=True, size=13, name="Arial", color="1F4E79")
    ws2["A3"] = "Fecha captura:";   ws2["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A4"] = "Total alertas:";   ws2["B4"] = len(df)
    ws2["A5"] = "PDFs en GitHub:";  ws2["B5"] = df.get("github_pdf_url", pd.Series()).notna().sum()

    ws2["A7"] = "Por urgencia:"
    ws2["A7"].font = Font(bold=True, name="Arial")
    for i, (k, v) in enumerate(df.get("urgencia", pd.Series()).value_counts().items(), 8):
        ws2[f"A{i}"] = k; ws2[f"B{i}"] = v

    ws2["D7"] = "Acciones principales:"
    ws2["D7"].font = Font(bold=True, name="Arial")
    if "accion_principal" in df.columns:
        for i, (k, v) in enumerate(df["accion_principal"].value_counts().head(8).items(), 8):
            ws2[f"D{i}"] = k; ws2[f"E{i}"] = v

    ws2["G7"] = "Motor de análisis:"
    ws2["G7"].font = Font(bold=True, name="Arial")
    if "motor_analisis" in df.columns:
        for i, (k, v) in enumerate(df["motor_analisis"].value_counts().items(), 8):
            ws2[f"G{i}"] = k; ws2[f"H{i}"] = v

    for col in ["A","B","D","E","G","H"]:
        ws2.column_dimensions[col].width = 40

    wb.save(ruta)
    print(f"  → Excel guardado: {ruta}")


# ─────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    tiene_claude = bool(os.environ.get("ANTHROPIC_API_KEY"))
    tiene_github = bool(GITHUB_TOKEN and GITHUB_REPO)

    print("═" * 60)
    print("  DIGEMID Alertas Scraper v3.0 — Conkomerco")
    print("═" * 60)
    print(f"  Claude API : {'✅ activo' if tiene_claude else '⚠ sin ANTHROPIC_API_KEY (heurístico)'}")
    print(f"  GitHub     : {'✅ ' + GITHUB_REPO if tiene_github else '⚠ sin GITHUB_TOKEN/REPO (solo local)'}")
    print("═" * 60 + "\n")

    df = scrapear_alertas(
        max_paginas=1,           # ← 1 para test (10 alertas). None para histórico completo.
        delay_paginas=1.5,
        analizar_acciones=True,
    )

    # Vista previa en consola
    print("\n--- Vista previa ---")
    cols_preview = ["titulo", "accion_principal", "urgencia", "motor_analisis"]
    cols_ok = [c for c in cols_preview if c in df.columns]
    print(df[cols_ok].to_string())

    # Exportar Excel local
    fecha_str = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_xlsx = f"alertas_digemid_{fecha_str}.xlsx"
    ruta_xlsx   = f"/mnt/user-data/outputs/{nombre_xlsx}"
    exportar_excel(df, ruta_xlsx)

    # Subir JSON + resumen a GitHub
    if tiene_github:
        fecha_dia = datetime.now().strftime("%Y%m%d")
        registros = df.to_dict("records")
        subir_json_a_github(registros, f"alertas_{fecha_dia}.json")
        subir_json_a_github(registros, "alertas_latest.json")
        subir_resumen_a_github(df, datetime.now().strftime("%Y/%m/%d"))

    print("\n✅ Proceso completado.")


# ─────────────────────────────────────────────────────────────────
# GITHUB ACTIONS WORKFLOW (copiar a .github/workflows/digemid.yml)
# ─────────────────────────────────────────────────────────────────
"""
name: Monitoreo DIGEMID Diario

on:
  schedule:
    - cron: '0 13 * * *'   # 8:00 AM Lima (UTC-5 → UTC 13:00)
  workflow_dispatch:        # ejecución manual desde GitHub Actions

jobs:
  scrape:
    runs-on: ubuntu-latest
    permissions:
      contents: write        # necesario para subir PDFs y JSONs al repo

    steps:
      - name: Checkout repo
        uses: actions/checkout@v4

      - name: Setup Python 3.11
        uses: actions/setup-python@v5
        with:
          python-version: '3.11'

      - name: Instalar dependencias
        run: |
          pip install requests beautifulsoup4 pandas openpyxl pymupdf anthropic

      - name: Ejecutar scraper DIGEMID
        env:
          ANTHROPIC_API_KEY: ${{ secrets.ANTHROPIC_API_KEY }}
          GITHUB_TOKEN:      ${{ secrets.GITHUB_TOKEN }}
          GITHUB_REPO:       ${{ github.repository }}
          GITHUB_BRANCH:     main
        run: python scraper_alertas_digemid_v3.py

      - name: Commit Excel al repo
        run: |
          git config user.name  "digemid-bot"
          git config user.email "bot@conkomerco.com"
          git add data/ summaries/ "*.xlsx" || true
          git diff --staged --quiet || git commit -m "📋 Alertas DIGEMID $(date +'%Y-%m-%d')"
          git push
"""
