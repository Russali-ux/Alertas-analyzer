#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FDA AEMS Monitor
================
Scraper de "New Safety Information or Potential Signals of Serious Risks"
del FDA Adverse Event Monitoring System (AEMS, antes FAERS).

Flujo:
  1. Lee la página índice (vigentes) y la de archivados.
  2. Descubre todos los reportes trimestrales enlazados (2008 -> actual).
  3. Descarga cada reporte trimestral y extrae su tabla de señales.
  4. Compara contra una "base_datos" local (JSON) para detectar:
       - reportes nuevos (trimestres no vistos antes)
       - filas nuevas dentro de un reporte ya visto
       - filas cuyo estatus cambió (ej: "en evaluación" -> "FDA determinó
         que no se requiere acción" / "cambio de etiquetado", etc.)
  5. Escribe:
       - fda_aems_data.json   (base de datos acumulada, fuente de verdad)
       - fda_aems_reporte.xlsx (reporte Excel con todas las señales + resumen)
       - fda_aems_cambios.json (solo lo detectado como nuevo/cambiado en esta corrida)

Uso:
    python fda_aems_monitor.py                # corrida normal
    python fda_aems_monitor.py --no-archived   # solo reportes vigentes (más rápido)
    python fda_aems_monitor.py --db ruta.json  # usar otra ruta de base de datos
"""

import argparse
import json
import re
import sys
import time
import hashlib
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import requests
from bs4 import BeautifulSoup

BASE = "https://www.fda.gov"
INDEX_URL = f"{BASE}/drugs/fda-adverse-event-monitoring-system-aems/new-safety-information-or-potential-signals-serious-risks-identified-fda-adverse-event-monitoring"
ARCHIVED_URL = f"{BASE}/drugs/fda-adverse-event-monitoring-system-aems/archived-quarterly-reports-new-safety-information-or-potential-signals-serious-risks-identified-fda"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": f"{BASE}/",
}

DEFAULT_DB_PATH = "fda_aems_data.json"
DEFAULT_XLSX_PATH = "fda_aems_reporte.xlsx"
DEFAULT_CHANGES_PATH = "fda_aems_cambios.json"

SESSION = requests.Session()
SESSION.headers.update(HEADERS)


# --------------------------------------------------------------------------- #
# Utilidades
# --------------------------------------------------------------------------- #

def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def get(url: str, retries: int = 3, timeout: int = 30):
    import random
    last_err = None
    for i in range(retries):
        try:
            r = SESSION.get(url, timeout=timeout)
            r.raise_for_status()
            return r
        except Exception as e:  # noqa: BLE001
            last_err = e
            wait = 5 * (i + 1) + random.uniform(1, 4)
            log(f"  reintento {i+1}/{retries} para {url} ({e}) — esperando {wait:.1f}s")
            time.sleep(wait)
    raise RuntimeError(f"No se pudo obtener {url}: {last_err}")


def normalize_ws(text: str) -> str:
    if text is None:
        return ""
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def make_row_id(quarter_slug: str, producto: str, senal: str) -> str:
    """ID estable para una fila (producto+señal dentro de un trimestre)."""
    raw = f"{quarter_slug}||{producto.strip().lower()}||{senal.strip().lower()}"
    raw = unicodedata.normalize("NFKD", raw)
    return hashlib.sha1(raw.encode("utf-8", "ignore")).hexdigest()[:16]


def slug_from_url(url: str) -> str:
    return url.rstrip("/").split("/")[-1]


# --------------------------------------------------------------------------- #
# Extracción de principio(s) activo(s) desde el campo "Producto"
# --------------------------------------------------------------------------- #
# El campo "Producto" de la FDA mezcla nombre comercial + forma farmacéutica +
# principio activo entre paréntesis/corchetes, y a veces lista varios productos
# de una misma clase terapéutica (ej. "SGLT2 inhibitors: Farxiga (dapagliflozin)
# tablet Jardiance (empagliflozin) tablet"). Esta función extrae solo los
# principios activos (contenido entre paréntesis/corchetes, filtrando ruido de
# forma farmacéutica/vía de administración/dosis), separados por " / ".
# Si no hay paréntesis, el "producto" ya suele ser el nombre genérico o una
# clase terapéutica (ej. "Acetaminophen", "Metformin-containing drug products")
# y se usa tal cual.

_NOISE_PATTERNS = [
    r'^several strengths?$', r'^with alcohol$', r'^numerous$', r'^transdermal system$',
    r'^rDNA origin$', r'^recombinant$', r'^\d+%?$', r'^various( strengths)?$',
    r'^for (subcutaneous|intravenous|iv|im|intramuscular|oral|topical) use$',
    r'^(subcutaneous|intravenous|iv|im|intramuscular|oral|topical)( use)?$',
    r'^extended.release$', r'^delayed.release$', r'^immediate.release$',
    r'^oral (solution|suspension)$', r'^injection$', r'^tablet[s]?$', r'^capsule[s]?$',
    r'^patch$', r'^spray$', r'^gel$', r'^cream$', r'^lotion$', r'^cartridges?$',
]
_NOISE_RE = re.compile('|'.join(_NOISE_PATTERNS), re.I)


def extract_active_ingredients(producto: str) -> str:
    """Devuelve solo el/los principio(s) activo(s) a partir del campo Producto."""
    if not producto:
        return ""
    matches = re.findall(r"[\(\[]([^\(\)\[\]]+)[\)\]]", producto)
    seen, seen_lower = [], set()
    for m in matches:
        m = normalize_ws(m)
        if not m or len(m) < 3:
            continue
        if _NOISE_RE.match(m):
            continue
        if m.lower() not in seen_lower:
            seen.append(m)
            seen_lower.add(m.lower())
    if seen:
        return " / ".join(seen)
    # Sin paréntesis: el producto ya suele ser el nombre genérico o la clase.
    return producto.strip()


# --------------------------------------------------------------------------- #
# Descubrimiento de reportes trimestrales
# --------------------------------------------------------------------------- #

QUARTER_LINK_RE = re.compile(
    r"(january-march|april-june|july-september|october-december)-(\d{4})", re.I
)


def discover_quarterly_links(html: str) -> list[dict]:
    """Extrae enlaces a reportes trimestrales de una página índice/archivo."""
    soup = BeautifulSoup(html, "html.parser")
    found = {}
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href.startswith("/drugs/fda-adverse-event-monitoring-system-aems/") and \
           not href.startswith("/drugs/fda-adverse-event-reporting-system-faers/"):
            continue
        m = QUARTER_LINK_RE.search(href)
        if not m:
            continue
        full_url = href if href.startswith("http") else BASE + href
        slug = slug_from_url(full_url)
        period_key = f"{m.group(1).lower()}-{m.group(2)}"
        found[slug] = {
            "url": full_url,
            "slug": slug,
            "period_raw": period_key,
            "year": int(m.group(2)),
            "label_text": normalize_ws(a.get_text()),
        }
    return list(found.values())


PERIOD_ORDER = {
    "january-march": 1,
    "april-june": 2,
    "july-september": 3,
    "october-december": 4,
}


def period_sort_key(entry: dict):
    q, y = entry["period_raw"].rsplit("-", 1)
    return (int(y), PERIOD_ORDER.get(q, 0))


def quarter_display_name(entry: dict) -> str:
    q = entry["period_raw"].rsplit("-", 1)[0]
    names = {
        "january-march": "Enero - Marzo",
        "april-june": "Abril - Junio",
        "july-september": "Julio - Septiembre",
        "october-december": "Octubre - Diciembre",
    }
    return f"{names.get(q, q)} {entry['year']}"


# --------------------------------------------------------------------------- #
# Parseo de un reporte trimestral individual
# --------------------------------------------------------------------------- #

def parse_quarterly_report(html: str, quarter_entry: dict) -> dict:
    soup = BeautifulSoup(html, "html.parser")

    # "Content current as of:" -> fecha de contenido
    content_date = None
    for li in soup.find_all(["li", "div", "p"]):
        txt = normalize_ws(li.get_text())
        if txt.lower().startswith("content current as of"):
            m = re.search(r"(\d{2}/\d{2}/\d{4})", txt)
            if m:
                content_date = m.group(1)
            break

    table = soup.find("table")
    rows_out = []
    if table:
        trs = table.find_all("tr")
        header_cells = [normalize_ws(c.get_text()) for c in trs[0].find_all(["th", "td"])]

        for tr in trs[1:]:
            cells = tr.find_all(["td", "th"])
            if not cells:
                continue
            values = []
            for c in cells:
                # separar líneas de <br> con " | " para no perder productos múltiples
                for br in c.find_all("br"):
                    br.replace_with("\n")
                txt = normalize_ws(c.get_text(separator="\n"))
                values.append(txt)
            if len(values) < 2:
                continue
            producto = values[0] if len(values) > 0 else ""
            senal = values[1] if len(values) > 1 else ""
            info_adicional = values[2] if len(values) > 2 else ""

            if not producto and not senal:
                continue

            row_id = make_row_id(quarter_entry["slug"], producto, senal)
            rows_out.append({
                "row_id": row_id,
                "producto": producto,
                "senal_riesgo": senal,
                "informacion_adicional": info_adicional,
            })

    return {
        "slug": quarter_entry["slug"],
        "url": quarter_entry["url"],
        "periodo": quarter_display_name(quarter_entry),
        "anio": quarter_entry["year"],
        "content_current_as_of": content_date,
        "n_filas": len(rows_out),
        "filas": rows_out,
        "scraped_at": datetime.now(timezone.utc).isoformat(),
    }


# --------------------------------------------------------------------------- #
# Base de datos local (JSON acumulado)
# --------------------------------------------------------------------------- #

def load_db(path: Path) -> dict:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:  # noqa: BLE001
            log(f"  No se pudo leer base de datos existente ({e}); se crea una nueva.")
    return {"reportes": {}, "last_run": None}


def save_db(db: dict, path: Path):
    path.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")


# --------------------------------------------------------------------------- #
# Detección de cambios
# --------------------------------------------------------------------------- #

def detect_changes(db_old: dict, reportes_nuevos: list[dict]) -> dict:
    cambios = {
        "reportes_nuevos": [],       # trimestres que no existían antes
        "filas_nuevas": [],          # filas nuevas en reportes ya conocidos
        "filas_modificadas": [],     # filas cuyo texto de info adicional cambió
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    reportes_old = db_old.get("reportes", {})

    for rep in reportes_nuevos:
        slug = rep["slug"]
        if slug not in reportes_old:
            cambios["reportes_nuevos"].append({
                "slug": slug,
                "periodo": rep["periodo"],
                "url": rep["url"],
                "n_filas": rep["n_filas"],
            })
            continue

        old_rows = {f["row_id"]: f for f in reportes_old[slug].get("filas", [])}
        for f in rep["filas"]:
            if f["row_id"] not in old_rows:
                cambios["filas_nuevas"].append({
                    "periodo": rep["periodo"],
                    "producto": f["producto"],
                    "senal_riesgo": f["senal_riesgo"],
                    "informacion_adicional": f["informacion_adicional"],
                })
            else:
                old_info = old_rows[f["row_id"]].get("informacion_adicional", "")
                if normalize_ws(old_info) != normalize_ws(f["informacion_adicional"]):
                    cambios["filas_modificadas"].append({
                        "periodo": rep["periodo"],
                        "producto": f["producto"],
                        "senal_riesgo": f["senal_riesgo"],
                        "info_anterior": old_info,
                        "info_nueva": f["informacion_adicional"],
                    })

    return cambios


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #

def build_excel(reportes: list[dict], cambios: dict, out_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    brand_fill = PatternFill("solid", fgColor="1E3A78")
    header_fill = PatternFill("solid", fgColor="E8ECF6")
    new_fill = PatternFill("solid", fgColor="DBEAFE")
    mod_fill = PatternFill("solid", fgColor="FEF9C3")
    white_bold = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    header_font = Font(name="Arial", bold=True, size=10)
    normal_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    # ---- Hoja 1: Todas las señales ----
    ws = wb.active
    ws.title = "Señales AEMS"

    ws.merge_cells("A1:F1")
    ws["A1"] = "ConkoSafe IA — FDA AEMS | Señales de Riesgo Serio (Farmacovigilancia)"
    ws["A1"].font = white_bold
    ws["A1"].fill = brand_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Reportes trimestrales: {len(reportes)}  |  Total señales: {sum(r['n_filas'] for r in reportes)}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9)

    headers = ["Periodo", "Año", "Producto (Trade / Ingrediente Activo)", "Principio(s) Activo(s)",
               "Señal de Riesgo Serio", "Información Adicional", "URL Fuente"]
    hr = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = wrap

    changed_new_ids = {(f["periodo"], f["producto"], f["senal_riesgo"]) for f in cambios.get("filas_nuevas", [])}
    changed_mod_ids = {(f["periodo"], f["producto"], f["senal_riesgo"]) for f in cambios.get("filas_modificadas", [])}

    r = hr + 1
    for rep in sorted(reportes, key=lambda x: (x["anio"], x["slug"]), reverse=True):
        for f in rep["filas"]:
            key = (rep["periodo"], f["producto"], f["senal_riesgo"])
            row_vals = [rep["periodo"], rep["anio"], f["producto"], extract_active_ingredients(f["producto"]),
                        f["senal_riesgo"], f["informacion_adicional"], rep["url"]]
            for i, v in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=i, value=v)
                c.font = normal_font
                c.border = border
                c.alignment = wrap
                if key in changed_new_ids:
                    c.fill = new_fill
                elif key in changed_mod_ids:
                    c.fill = mod_fill
            r += 1

    widths = [16, 7, 32, 26, 26, 38, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ---- Hoja 2: Cambios detectados en esta corrida ----
    ws2 = wb.create_sheet("Cambios Detectados")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Cambios detectados en esta corrida"
    ws2["A1"].font = white_bold
    ws2["A1"].fill = brand_fill
    ws2.row_dimensions[1].height = 22

    row = 3
    ws2.cell(row=row, column=1, value="Reportes trimestrales nuevos").font = Font(bold=True, size=11)
    row += 1
    for h, w in zip(["Periodo", "N° señales", "URL"], [16, 10, 60]):
        pass
    headers2 = ["Periodo", "N° señales", "URL"]
    for i, h in enumerate(headers2, start=1):
        c = ws2.cell(row=row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1
    if cambios["reportes_nuevos"]:
        for rn in cambios["reportes_nuevos"]:
            ws2.cell(row=row, column=1, value=rn["periodo"])
            ws2.cell(row=row, column=2, value=rn["n_filas"])
            ws2.cell(row=row, column=3, value=rn["url"])
            row += 1
    else:
        ws2.cell(row=row, column=1, value="(ninguno)")
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="Señales nuevas (en reportes ya conocidos)").font = Font(bold=True, size=11)
    row += 1
    headers3 = ["Periodo", "Producto", "Señal de Riesgo", "Info Adicional"]
    for i, h in enumerate(headers3, start=1):
        c = ws2.cell(row=row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1
    if cambios["filas_nuevas"]:
        for f in cambios["filas_nuevas"]:
            ws2.cell(row=row, column=1, value=f["periodo"])
            ws2.cell(row=row, column=2, value=f["producto"])
            ws2.cell(row=row, column=3, value=f["senal_riesgo"])
            ws2.cell(row=row, column=4, value=f["informacion_adicional"])
            for c in range(1, 5):
                ws2.cell(row=row, column=c).alignment = wrap
            row += 1
    else:
        ws2.cell(row=row, column=1, value="(ninguna)")
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="Señales con estatus / info adicional modificada").font = Font(bold=True, size=11)
    row += 1
    headers4 = ["Periodo", "Producto", "Señal de Riesgo", "Info Anterior", "Info Nueva"]
    for i, h in enumerate(headers4, start=1):
        c = ws2.cell(row=row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1
    if cambios["filas_modificadas"]:
        for f in cambios["filas_modificadas"]:
            ws2.cell(row=row, column=1, value=f["periodo"])
            ws2.cell(row=row, column=2, value=f["producto"])
            ws2.cell(row=row, column=3, value=f["senal_riesgo"])
            ws2.cell(row=row, column=4, value=f["info_anterior"])
            ws2.cell(row=row, column=5, value=f["info_nueva"])
            for c in range(1, 6):
                ws2.cell(row=row, column=c).alignment = wrap
            row += 1
    else:
        ws2.cell(row=row, column=1, value="(ninguna)")
        row += 1

    for col, w in zip("ABCDEF", [18, 30, 26, 38, 38]):
        ws2.column_dimensions[col].width = w

    # ---- Hoja 3: Resumen por trimestre ----
    ws3 = wb.create_sheet("Resumen por Trimestre")
    headers5 = ["Periodo", "Año", "N° Señales", "Content Current As Of", "URL"]
    for i, h in enumerate(headers5, start=1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
    r = 2
    for rep in sorted(reportes, key=lambda x: (x["anio"], x["slug"]), reverse=True):
        ws3.cell(row=r, column=1, value=rep["periodo"])
        ws3.cell(row=r, column=2, value=rep["anio"])
        ws3.cell(row=r, column=3, value=rep["n_filas"])
        ws3.cell(row=r, column=4, value=rep.get("content_current_as_of") or "")
        ws3.cell(row=r, column=5, value=rep["url"])
        r += 1
    for col, w in zip("ABCDE", [16, 8, 12, 20, 60]):
        ws3.column_dimensions[col].width = w

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Resumen para correo
# --------------------------------------------------------------------------- #

def build_email_summary(cambios: dict, total_reportes: int, total_senales: int) -> str:
    n_rep_nuevos = len(cambios["reportes_nuevos"])
    n_filas_nuevas = len(cambios["filas_nuevas"])
    n_filas_mod = len(cambios["filas_modificadas"])
    n_total_cambios = n_rep_nuevos + n_filas_nuevas + n_filas_mod

    lines = []
    lines.append("ConkoSafe IA — Monitor FDA AEMS (Señales de Riesgo Serio)")
    lines.append(f"Corrida: {datetime.now().strftime('%Y-%m-%d %H:%M')} (hora del runner)")
    lines.append("")
    lines.append(f"Base acumulada: {total_reportes} reportes trimestrales | {total_senales} señales totales")
    lines.append("")

    if n_total_cambios == 0:
        lines.append("No se detectaron cambios respecto a la corrida anterior.")
        lines.append("Se adjunta el reporte Excel actualizado de todas formas para referencia.")
        return "\n".join(lines)

    lines.append(f"Cambios detectados en esta corrida: {n_total_cambios}")
    lines.append("")

    if n_rep_nuevos:
        lines.append(f"Reportes trimestrales nuevos ({n_rep_nuevos}):")
        for rn in cambios["reportes_nuevos"]:
            lines.append(f"  - {rn['periodo']} ({rn['n_filas']} señales) — {rn['url']}")
        lines.append("")

    if n_filas_nuevas:
        lines.append(f"Señales nuevas en reportes ya conocidos ({n_filas_nuevas}):")
        for f in cambios["filas_nuevas"]:
            lines.append(f"  - [{f['periodo']}] {f['producto']} — {f['senal_riesgo']}")
            lines.append(f"      {f['informacion_adicional']}")
        lines.append("")

    if n_filas_mod:
        lines.append(f"Señales con estatus/información actualizada ({n_filas_mod}):")
        for f in cambios["filas_modificadas"]:
            lines.append(f"  - [{f['periodo']}] {f['producto']} — {f['senal_riesgo']}")
            lines.append(f"      Antes: {f['info_anterior']}")
            lines.append(f"      Ahora: {f['info_nueva']}")
        lines.append("")

    lines.append("Detalle completo en el Excel adjunto (hoja 'Cambios Detectados').")
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    ap = argparse.ArgumentParser(description="FDA AEMS monitor / scraper")
    ap.add_argument("--db", default=DEFAULT_DB_PATH, help="Ruta al JSON de base de datos acumulada")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX_PATH, help="Ruta de salida del Excel")
    ap.add_argument("--changes", default=DEFAULT_CHANGES_PATH, help="Ruta de salida del JSON de cambios de esta corrida")
    ap.add_argument("--resumen", default=None, help="Ruta de salida de un .txt con el resumen para el borrador de correo")
    ap.add_argument("--no-archived", action="store_true", help="No incluir la página de reportes archivados")
    ap.add_argument("--sleep", type=float, default=0.8, help="Segundos de espera entre requests")
    args = ap.parse_args()

    db_path = Path(args.db)
    xlsx_path = Path(args.xlsx)
    changes_path = Path(args.changes)

    log("Descargando página índice (reportes vigentes)...")
    index_html = get(INDEX_URL).text
    links = discover_quarterly_links(index_html)

    if not args.no_archived:
        log("Descargando página de reportes archivados...")
        try:
            archived_html = get(ARCHIVED_URL).text
            links += discover_quarterly_links(archived_html)
        except Exception as e:  # noqa: BLE001
            log(f"  Aviso: no se pudo leer la página de archivados ({e}). Se continúa solo con vigentes.")

    # deduplicar por slug
    dedup = {}
    for l in links:
        dedup[l["slug"]] = l
    links = sorted(dedup.values(), key=period_sort_key)

    log(f"Se detectaron {len(links)} reportes trimestrales enlazados (2008-actual, según disponibilidad).")

    db = load_db(db_path)

    reportes_nuevos_completos = []
    for i, entry in enumerate(links, start=1):
        log(f"[{i}/{len(links)}] {quarter_display_name(entry)} -> {entry['slug']}")
        try:
            html = get(entry["url"]).text
        except Exception as e:  # noqa: BLE001
            log(f"  ERROR descargando {entry['url']}: {e}")
            continue
        parsed = parse_quarterly_report(html, entry)
        reportes_nuevos_completos.append(parsed)
        time.sleep(args.sleep)

    cambios = detect_changes(db, reportes_nuevos_completos)

    # actualizar base de datos
    for rep in reportes_nuevos_completos:
        db["reportes"][rep["slug"]] = rep
    db["last_run"] = datetime.now(timezone.utc).isoformat()
    save_db(db, db_path)

    changes_path.write_text(json.dumps(cambios, ensure_ascii=False, indent=2), encoding="utf-8")

    all_reportes = list(db["reportes"].values())
    log(f"Generando Excel con {len(all_reportes)} reportes trimestrales...")
    build_excel(all_reportes, cambios, xlsx_path)

    total_senales = sum(r["n_filas"] for r in all_reportes)

    if args.resumen:
        resumen_path = Path(args.resumen)
        resumen_texto = build_email_summary(cambios, len(all_reportes), total_senales)
        resumen_path.write_text(resumen_texto, encoding="utf-8")

    log("=" * 60)
    log(f"Base de datos:      {db_path}  ({len(all_reportes)} reportes, {total_senales} señales)")
    log(f"Excel:              {xlsx_path}")
    log(f"Cambios esta corrida: {changes_path}")
    if args.resumen:
        log(f"Resumen p/correo:   {args.resumen}")
    log(f"  Reportes nuevos:      {len(cambios['reportes_nuevos'])}")
    log(f"  Señales nuevas:       {len(cambios['filas_nuevas'])}")
    log(f"  Señales modificadas:  {len(cambios['filas_modificadas'])}")
    log("=" * 60)


if __name__ == "__main__":
    sys.exit(main())
