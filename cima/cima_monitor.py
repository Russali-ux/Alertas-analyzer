#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ConkoSafe IA — Monitor CIMA (AEMPS): Ficha Técnica y Prospecto.
Runner SELF-CONTAINED para GitHub Actions (Russali-ux/Alertas-analyzer).

Módulo totalmente independiente del monitor DIGEMID del repo. NO usa Google Drive.

Pipeline por corrida:
  1. Descarga registroCambios de CIMA vía API REST (últimos N días, default 30).
  2. Filtra medicamentos con cambio en Ficha Técnica y/o Prospecto.
  3. Genera, codificado por fecha, dentro de cima/data/ y cima/summaries/:
        - Excel  ConkosafeIA_Regulatorio_YYYYMMDD.xlsx   (5 hojas)
        - JSON   cima_YYYYMMDD.json                       (datos para el visor web)
        - JSON   cima_latest.json                         (copia del último run)
        - JSON   index.json                               (índice de todas las fechas)
        - MD     cima_YYYY-MM-DD.md                       (resumen legible)

Uso:
    python cima_monitor.py                 # últimos 30 días
    python cima_monitor.py --dias 60       # últimos 60 días
    python cima_monitor.py --desde 01/05/2026 --hasta 31/05/2026
    python cima_monitor.py --data-dir cima/data --summary-dir cima/summaries

Fuente oficial y pública:
    https://cima.aemps.es/cima/rest/registroCambios
"""

import argparse
import glob
import json
import os
import re
import sys
import time
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────── CONSTANTES ────────────────────────────────────────

CIMA_API_BASE    = "https://cima.aemps.es/cima/rest"
CIMA_CAMBIOS_URL = f"{CIMA_API_BASE}/registroCambios"
CIMA_PDF_FT      = "https://cima.aemps.es/cima/pdfs/ft/{nr}/FT_{nr}.pdf"
CIMA_HTML_FT     = "https://cima.aemps.es/cima/dochtml/ft/{nr}/FT_{nr}.html"
CIMA_PDF_P       = "https://cima.aemps.es/cima/pdfs/p/{nr}/P_{nr}.pdf"
CIMA_HTML_P      = "https://cima.aemps.es/cima/dochtml/p/{nr}/P_{nr}.html"

TIMEOUT    = 30
RETRY_MAX  = 3
RETRY_WAIT = 2.0
PAGE_DELAY = 0.15

TIPO_CAMBIO = {1: "NUEVO", 2: "BAJA", 3: "MODIFICADO"}
CAMBIO_DESC = {
    "estado":         "Estado de autorización",
    "comerc":         "Estado de comercialización",
    "prosp":          "Prospecto",
    "ft":             "Ficha Técnica",
    "psum":           "Problemas de suministro",
    "notasSeguridad": "Notas de seguridad",
    "matinf":         "Materiales informativos",
    "otros":          "Otros cambios",
}

# ──────────────────────── ESTILOS EXCEL ─────────────────────────────────────

COLOR_HEADER = "1A3A5C"
COLOR_FT     = "D5E8D4"
COLOR_PROSP  = "DAE8FC"
COLOR_AMBOS  = "FFE6CC"
COLOR_ALT    = "F8F8F8"

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS_BASE = [
    "Nº Registro", "Nombre del Medicamento", "Laboratorio Titular",
    "Fecha del Cambio", "Tipo de Cambio", "Otros Cambios",
    "Ficha Técnica\n(PDF)", "Ficha Técnica\n(HTML)",
    "Prospecto\n(PDF)", "Prospecto\n(HTML)",
]
ANCHOS = [16, 45, 30, 18, 13, 40, 18, 18, 18, 18]


# ──────────────────────── UTILIDADES ────────────────────────────────────────

def log(msg, level="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = {"INFO": "[i] ", "OK": "[OK] ", "WARN": "[!] ", "ERR": "[X] "}.get(level, "")
    print(f"[{ts}] {prefix}{msg}", file=sys.stderr)


def retry_get(url, headers=None, timeout=TIMEOUT):
    for attempt in range(1, RETRY_MAX + 1):
        try:
            r = requests.get(url, headers=headers, timeout=timeout)
            r.raise_for_status()
            return r
        except Exception:
            if attempt == RETRY_MAX:
                raise
            time.sleep(RETRY_WAIT * attempt)


def epoch_to_dt(ms):
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc)


def epoch_to_str(ms):
    return epoch_to_dt(ms).strftime("%d/%m/%Y %H:%M")


def build_urls(nr):
    nr = str(nr).strip()
    return (
        CIMA_PDF_FT.format(nr=nr),
        CIMA_HTML_FT.format(nr=nr),
        CIMA_PDF_P.format(nr=nr),
        CIMA_HTML_P.format(nr=nr),
    )


def classify(flags):
    has_ft    = "ft"    in flags
    has_prosp = "prosp" in flags
    if has_ft and has_prosp: return "both"
    if has_ft:               return "ft"
    if has_prosp:            return "prosp"
    return "none"


# ──────────────────────── 1. DESCARGA CIMA ──────────────────────────────────

def fetch_registro_cambios(fecha_desde, fecha_hasta=None):
    records = []
    pagina  = 1
    total   = None
    log(f"Descargando registroCambios desde {fecha_desde}...")

    while True:
        url  = f"{CIMA_CAMBIOS_URL}?fecha={fecha_desde}&pagina={pagina}"
        data = retry_get(url).json()

        if isinstance(data, dict) and "error" in data:
            raise RuntimeError(f"CIMA devolvió error: {data['error']}")

        total      = data.get("totalFilas", 0)
        resultados = data.get("resultados", [])
        if not resultados:
            break

        records.extend(resultados)
        log(f"  página {pagina} — {len(records)}/{total} registros")

        if len(records) >= total:
            break
        pagina += 1
        time.sleep(PAGE_DELAY)

    if fecha_hasta is not None:
        records = [r for r in records if epoch_to_dt(r["fecha"]) <= fecha_hasta]

    log(f"Total tras filtros de fecha: {len(records)} registros", "OK")
    return records


# ──────────────────────── 2. PROCESAMIENTO ──────────────────────────────────

def process_records(raw_records):
    """Deduplica y filtra solo FT/Prospecto, enriquece con URLs y metadatos."""
    seen     = set()
    enriched = []

    for r in raw_records:
        key = (r.get("nregistro"), r.get("fecha"))
        if key in seen:
            continue
        seen.add(key)

        flags = r.get("cambio", []) or []
        cat   = classify(flags)
        if cat == "none":
            continue

        nreg = r.get("nregistro", "")
        ft_pdf, ft_html, p_pdf, p_html = build_urls(nreg)
        has_ft    = "ft"    in flags
        has_prosp = "prosp" in flags

        otros = " | ".join(
            CAMBIO_DESC.get(c, c) for c in flags if c not in ("ft", "prosp")
        )

        enriched.append({
            "nreg":        str(nreg),
            "nombre":      r.get("nombre", ""),
            "lab":         r.get("labtitular", ""),
            "fecha_epoch": r["fecha"],
            "fecha":       epoch_to_str(r["fecha"]),
            "tipo":        TIPO_CAMBIO.get(r.get("tipoCambio"), str(r.get("tipoCambio"))),
            "otros":       otros,
            "has_ft":      has_ft,
            "has_prosp":   has_prosp,
            "cat":         cat,
            "ft_pdf":      ft_pdf   if has_ft    else None,
            "ft_html":     ft_html  if has_ft    else None,
            "p_pdf":       p_pdf    if has_prosp else None,
            "p_html":      p_html   if has_prosp else None,
        })

    enriched.sort(key=lambda x: x["fecha_epoch"], reverse=True)
    log(f"Registros con FT y/o Prospecto: {len(enriched)}", "OK")
    return enriched


# ──────────────────────── 3. GENERACIÓN EXCEL ───────────────────────────────

def apply_row(ws, ri, vals, fill, link_cols=None):
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill  = fill
        c.border = BORDER
        c.font  = Font(name="Arial", size=9)
        is_center = ci in (1, 4, 5) or ci >= 7
        c.alignment = Alignment(
            horizontal="center" if is_center else "left",
            vertical="center", wrap_text=False
        )
        if link_cols and ci in link_cols and link_cols[ci]:
            c.hyperlink = link_cols[ci]
            c.font = Font(name="Arial", size=9, color="0563C1", underline="single")


def make_header(ws, cols, row):
    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor=COLOR_HEADER)
    ha    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 36
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=col)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = BORDER


def fill_sheet(ws, data, titulo):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{get_column_letter(len(COLS_BASE))}1")
    t = ws.cell(row=1, column=1, value=titulo)
    t.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.cell(
        row=2, column=1,
        value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Registros: {len(data)}"
    ).font = Font(name="Arial", italic=True, size=9, color="888888")

    make_header(ws, COLS_BASE, row=3)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS_BASE))}3"

    fill_ft    = PatternFill("solid", fgColor=COLOR_FT)
    fill_prosp = PatternFill("solid", fgColor=COLOR_PROSP)
    fill_ambos = PatternFill("solid", fgColor=COLOR_AMBOS)
    fill_alt   = PatternFill("solid", fgColor=COLOR_ALT)
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    for ri, rec in enumerate(data, 4):
        row_fill = (fill_ambos if rec["cat"] == "both"  else
                    fill_ft    if rec["cat"] == "ft"    else
                    fill_prosp if rec["cat"] == "prosp" else
                    fill_alt   if ri % 2 == 0           else fill_white)

        vals = [
            rec["nreg"], rec["nombre"], rec["lab"], rec["fecha"], rec["tipo"], rec["otros"],
            "Ver PDF"  if rec["has_ft"]    else "",
            "Ver HTML" if rec["has_ft"]    else "",
            "Ver PDF"  if rec["has_prosp"] else "",
            "Ver HTML" if rec["has_prosp"] else "",
        ]
        apply_row(ws, ri, vals, row_fill, {
            7: rec["ft_pdf"], 8: rec["ft_html"],
            9: rec["p_pdf"],  10: rec["p_html"],
        })
        ws.row_dimensions[ri].height = 16

    for i, w in enumerate(ANCHOS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_resumen_sheet(ws, n_ft, n_prosp, n_ambos, total, periodo_label, top_labs):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="RESUMEN — Ficha Técnica & Prospecto (CIMA / AEMPS)")
    t.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1,
            value=f"Periodo analizado: {periodo_label}").font = Font(
                name="Arial", italic=True, size=9, color="888888")

    rows = [
        ("Categoría", "Cantidad", "% del Total", "Color"),
        ("Solo Ficha Técnica",     n_ft,    (n_ft/total*100    if total else 0), "Verde"),
        ("Solo Prospecto",         n_prosp, (n_prosp/total*100 if total else 0), "Azul"),
        ("FT + Prospecto (ambos)", n_ambos, (n_ambos/total*100 if total else 0), "Naranja"),
        ("TOTAL con FT o Prospecto", total, 100.0 if total else 0, ""),
    ]
    fills = [None, COLOR_FT, COLOR_PROSP, COLOR_AMBOS, "E8EEF5"]
    start = 4
    for i, row in enumerate(rows):
        ri = start + i
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci)
            if i == 0:
                c.value = val
                c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                c.fill  = PatternFill("solid", fgColor=COLOR_HEADER)
            else:
                if ci == 3 and isinstance(val, (int, float)):
                    c.value = f"{val:.1f}%"
                else:
                    c.value = val
                c.font = Font(name="Arial", size=10, bold=(i == 4))
                if fills[i]:
                    c.fill = PatternFill("solid", fgColor=fills[i])
            c.border    = BORDER
            c.alignment = Alignment(horizontal="center" if ci != 1 else "left",
                                    vertical="center")

    # Top laboratorios
    tl_start = start + len(rows) + 2
    ws.cell(row=tl_start, column=1, value="Top 10 laboratorios con más cambios").font = \
        Font(name="Arial", bold=True, size=11, color=COLOR_HEADER)
    ws.cell(row=tl_start + 1, column=1, value="Laboratorio").font = \
        Font(name="Arial", bold=True, color="FFFFFF")
    ws.cell(row=tl_start + 1, column=1).fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws.cell(row=tl_start + 1, column=2, value="Cambios").font = \
        Font(name="Arial", bold=True, color="FFFFFF")
    ws.cell(row=tl_start + 1, column=2).fill = PatternFill("solid", fgColor=COLOR_HEADER)
    for j, (lab, n) in enumerate(top_labs, 1):
        ws.cell(row=tl_start + 1 + j, column=1, value=lab).font = Font(name="Arial", size=10)
        ws.cell(row=tl_start + 1 + j, column=2, value=n).font = Font(name="Arial", size=10)
        ws.cell(row=tl_start + 1 + j, column=2).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14


def build_excel(data, periodo_label, out_path):
    wb = openpyxl.Workbook()

    ft_only  = [r for r in data if r["cat"] == "ft"]
    p_only   = [r for r in data if r["cat"] == "prosp"]
    ambos    = [r for r in data if r["cat"] == "both"]

    ws1 = wb.active
    ws1.title = "FT y Prospecto"
    fill_sheet(ws1, data, "TODOS LOS CAMBIOS — Ficha Técnica y/o Prospecto")

    fill_sheet(wb.create_sheet("Solo Ficha Técnica"), ft_only, "SOLO FICHA TÉCNICA")
    fill_sheet(wb.create_sheet("Solo Prospecto"),     p_only,  "SOLO PROSPECTO")
    fill_sheet(wb.create_sheet("FT + Prospecto"),     ambos,   "FT + PROSPECTO (ambos)")

    top_labs = Counter(r["lab"] for r in data if r["lab"]).most_common(10)
    build_resumen_sheet(wb.create_sheet("Resumen"),
                        len(ft_only), len(p_only), len(ambos), len(data),
                        periodo_label, top_labs)

    wb.save(out_path)
    log(f"Excel guardado: {out_path}", "OK")
    return {
        "total": len(data), "solo_ft": len(ft_only),
        "solo_prosp": len(p_only), "ambos": len(ambos),
        "top_labs": top_labs,
    }


# ──────────────────────── 4. JSON + MARKDOWN ────────────────────────────────

def build_json(data, stats, periodo, fecha_codigo, generado, excel_name, md_name):
    registros = [{
        "nreg": r["nreg"], "nombre": r["nombre"], "lab": r["lab"],
        "fecha": r["fecha"], "tipo": r["tipo"], "otros": r["otros"],
        "cat": r["cat"], "has_ft": r["has_ft"], "has_prosp": r["has_prosp"],
        "ft_pdf": r["ft_pdf"], "ft_html": r["ft_html"],
        "p_pdf": r["p_pdf"], "p_html": r["p_html"],
    } for r in data]

    return {
        "generado": generado.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "generado_str": generado.strftime("%d/%m/%Y %H:%M UTC"),
        "fecha_codigo": fecha_codigo,
        "periodo": periodo,
        "excel": excel_name,
        "markdown": md_name,
        "stats": {
            "total": stats["total"], "solo_ft": stats["solo_ft"],
            "solo_prosp": stats["solo_prosp"], "ambos": stats["ambos"],
            "top_labs": [{"lab": l, "n": n} for l, n in stats["top_labs"]],
        },
        "registros": registros,
    }


def build_markdown(data, stats, periodo, generado, excel_name):
    lines = []
    lines.append(f"# ConkoSafe IA — Cambios en Ficha Técnica y Prospecto (CIMA / AEMPS)")
    lines.append("")
    lines.append(f"**Generado:** {generado.strftime('%d/%m/%Y %H:%M UTC')}  ")
    lines.append(f"**Periodo:** {periodo['desde']} → {periodo['hasta']} "
                 f"(últimos {periodo['dias']} días)  ")
    lines.append(f"**Fuente:** [CIMA — AEMPS (registroCambios)]"
                 f"(https://cima.aemps.es/cima/rest/registroCambios)")
    lines.append("")
    lines.append("## Resumen")
    lines.append("")
    lines.append("| Categoría | Cantidad |")
    lines.append("|---|---:|")
    lines.append(f"| Solo Ficha Técnica | {stats['solo_ft']} |")
    lines.append(f"| Solo Prospecto | {stats['solo_prosp']} |")
    lines.append(f"| FT + Prospecto (ambos) | {stats['ambos']} |")
    lines.append(f"| **TOTAL con FT o Prospecto** | **{stats['total']}** |")
    lines.append("")
    lines.append(f"Descarga del detalle completo: `{excel_name}`")
    lines.append("")

    if stats["top_labs"]:
        lines.append("### Top laboratorios")
        lines.append("")
        for lab, n in stats["top_labs"]:
            lines.append(f"- {lab} — {n} cambio(s)")
        lines.append("")

    lines.append("## Detalle de cambios")
    lines.append("")
    if not data:
        lines.append("_Sin cambios en Ficha Técnica ni Prospecto en el periodo._")
    else:
        lines.append("| Nº Registro | Medicamento | Laboratorio | Fecha | Tipo | Cambio | Enlaces |")
        lines.append("|---|---|---|---|---|---|---|")
        cat_label = {"ft": "FT", "prosp": "Prospecto", "both": "FT + Prospecto"}
        for r in data:
            enlaces = []
            if r["has_ft"]:
                enlaces.append(f"[FT-PDF]({r['ft_pdf']})")
                enlaces.append(f"[FT-HTML]({r['ft_html']})")
            if r["has_prosp"]:
                enlaces.append(f"[P-PDF]({r['p_pdf']})")
                enlaces.append(f"[P-HTML]({r['p_html']})")
            nombre = (r["nombre"] or "").replace("|", "\\|")
            lab    = (r["lab"] or "").replace("|", "\\|")
            lines.append(
                f"| {r['nreg']} | {nombre} | {lab} | {r['fecha']} | {r['tipo']} "
                f"| {cat_label.get(r['cat'], r['cat'])} | {' · '.join(enlaces)} |"
            )
    lines.append("")
    lines.append("---")
    lines.append("_Nota: la API de CIMA indica que el documento completo (FT/Prospecto) "
                 "fue modificado, no la sección específica (4.1, 4.8, etc.)._")
    lines.append("")
    return "\n".join(lines)


def rebuild_index(data_dir):
    """Reconstruye index.json escaneando todos los cima_YYYYMMDD.json presentes."""
    entries = []
    for path in sorted(glob.glob(os.path.join(data_dir, "cima_*.json"))):
        base = os.path.basename(path)
        if base in ("cima_latest.json",):
            continue
        m = re.match(r"cima_(\d{8})\.json$", base)
        if not m:
            continue
        codigo = m.group(1)
        try:
            with open(path, encoding="utf-8") as f:
                d = json.load(f)
            total = d.get("stats", {}).get("total", 0)
            fecha_iso = f"{codigo[:4]}-{codigo[4:6]}-{codigo[6:8]}"
            entries.append({
                "codigo": codigo, "fecha": fecha_iso,
                "total": total, "archivo": base,
                "excel": d.get("excel", ""),
            })
        except Exception as e:
            log(f"No se pudo leer {base}: {e}", "WARN")

    entries.sort(key=lambda x: x["codigo"], reverse=True)
    index = {
        "actualizado": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "latest": entries[0]["archivo"] if entries else None,
        "fechas": entries,
    }
    with open(os.path.join(data_dir, "index.json"), "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    log(f"index.json reconstruido con {len(entries)} fecha(s)", "OK")


# ──────────────────────── MAIN ──────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Monitor CIMA FT/Prospecto (self-contained).")
    p.add_argument("--dias", type=int, default=30,
                   help="Ventana de días hacia atrás (default 30).")
    p.add_argument("--desde", help="Fecha desde dd/mm/yyyy (anula --dias).")
    p.add_argument("--hasta", help="Fecha hasta dd/mm/yyyy (opcional).")
    p.add_argument("--data-dir", default="cima/data",
                   help="Carpeta de salida para Excel/JSON (default cima/data).")
    p.add_argument("--summary-dir", default="cima/summaries",
                   help="Carpeta de salida para MD (default cima/summaries).")
    return p.parse_args()


def main():
    args = parse_args()

    hoy = datetime.now(timezone.utc)
    if args.desde:
        fecha_desde = args.desde
        desde_dt = datetime.strptime(args.desde, "%d/%m/%Y")
    else:
        desde_dt = hoy - timedelta(days=args.dias)
        fecha_desde = desde_dt.strftime("%d/%m/%Y")

    fecha_hasta_dt = None
    if args.hasta:
        fecha_hasta_dt = datetime.strptime(args.hasta, "%d/%m/%Y").replace(
            hour=23, minute=59, tzinfo=timezone.utc)
        fecha_hasta_label = args.hasta
    else:
        fecha_hasta_label = hoy.strftime("%d/%m/%Y")

    dias = (args.dias if not args.desde
            else (datetime.strptime(fecha_hasta_label, "%d/%m/%Y") - desde_dt).days)

    data_dir    = args.data_dir
    summary_dir = args.summary_dir
    Path(data_dir).mkdir(parents=True, exist_ok=True)
    Path(summary_dir).mkdir(parents=True, exist_ok=True)

    raw  = fetch_registro_cambios(fecha_desde, fecha_hasta_dt)
    data = process_records(raw)

    fecha_codigo = hoy.strftime("%Y%m%d")
    fecha_iso    = hoy.strftime("%Y-%m-%d")
    periodo_label = f"{fecha_desde} → {fecha_hasta_label} (últimos {dias} días)"
    periodo = {"desde": fecha_desde, "hasta": fecha_hasta_label, "dias": dias}

    excel_name = f"ConkosafeIA_Regulatorio_{fecha_codigo}.xlsx"
    md_name    = f"cima_{fecha_iso}.md"
    json_name  = f"cima_{fecha_codigo}.json"

    excel_path = os.path.join(data_dir, excel_name)
    stats = build_excel(data, periodo_label, excel_path)

    payload = build_json(data, stats, periodo, fecha_codigo, hoy,
                         excel_name, f"../summaries/{md_name}")
    with open(os.path.join(data_dir, json_name), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    with open(os.path.join(data_dir, "cima_latest.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    log(f"JSON guardado: {json_name} (+ cima_latest.json)", "OK")

    md = build_markdown(data, stats, periodo, hoy, excel_name)
    with open(os.path.join(summary_dir, md_name), "w", encoding="utf-8") as f:
        f.write(md)
    log(f"Markdown guardado: {md_name}", "OK")

    rebuild_index(data_dir)

    print(f"\n===== RESUMEN CIMA =====")
    print(f"Periodo:        {periodo_label}")
    print(f"Solo FT:        {stats['solo_ft']}")
    print(f"Solo Prospecto: {stats['solo_prosp']}")
    print(f"Ambos:          {stats['ambos']}")
    print(f"TOTAL:          {stats['total']}")
    print(f"Excel:          {excel_path}")
    print(f"========================")


if __name__ == "__main__":
    main()
